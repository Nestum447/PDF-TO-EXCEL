import streamlit as st
import tempfile
import os
import cv2
import pytesseract

from img2table.document import Image
from img2table.ocr import TesseractOCR

from openpyxl import Workbook

st.set_page_config(page_title="PDF / Imagen → Excel")

st.title("PDF o Imagen → Excel (manteniendo layout)")

archivo = st.file_uploader(
    "Sube PDF o imagen",
    type=["png","jpg","jpeg","pdf"]
)

if archivo is not None:

    extension = os.path.splitext(archivo.name)[1]

    with tempfile.NamedTemporaryFile(delete=False, suffix=extension) as tmp:

        tmp.write(archivo.read())
        ruta = tmp.name

    st.success("Archivo cargado")

    try:

        img = cv2.imread(ruta)

        if img is None:
            st.error("No se pudo leer la imagen")
            st.stop()

        data = pytesseract.image_to_data(
            img,
            output_type=pytesseract.Output.DICT
        )

        bloques_texto = []

        for i in range(len(data["text"])):

            texto = data["text"][i].strip()

            if texto != "":
                bloques_texto.append({
                    "text": texto,
                    "y": data["top"][i]
                })

        ocr = TesseractOCR(lang="eng")

        doc = Image(ruta)

        tables = doc.extract_tables(ocr=ocr)

        elementos = []

        for t in bloques_texto:

            elementos.append({
                "type": "text",
                "y": t["y"],
                "content": t["text"]
            })

        for table in tables:

            elementos.append({
                "type": "table",
                "y": table.bbox.y1,
                "content": table.df
            })

        elementos.sort(key=lambda x: x["y"])

        wb = Workbook()
        ws = wb.active

        fila = 1

        for el in elementos:

            if el["type"] == "text":

                ws.cell(row=fila, column=1).value = el["content"]
                fila += 1

            if el["type"] == "table":

                df = el["content"]

                for r in df.values:

                    for c, val in enumerate(r):

                        ws.cell(row=fila, column=c+1).value = val

                    fila += 1

                fila += 1

        salida = "resultado.xlsx"

        wb.save(salida)

        st.success("Excel generado")

        with open(salida,"rb") as f:

            st.download_button(
                "Descargar Excel",
                f,
                file_name="resultado.xlsx"
            )

    except Exception as e:

        st.error("Error procesando documento")
        st.text(str(e))

    finally:

        if os.path.exists(ruta):
            os.remove(ruta)
